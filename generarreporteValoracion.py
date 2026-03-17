from datetime import datetime
import pyodbc
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from io import BytesIO
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, PatternFill
import asyncio
from flask import jsonify
from app.config import get_connection
from app.utils import copiar_formato, Unirceldas, ConvertirNoneto0 , copiar
from app.Reportes.queryreportes import queryValorizacion
import os
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
#cnxn = Config.cnxn
#cursor = cnxn.cursor()

def GenerarReporteValorizacion(CodSed,NroOrden=""):
    cnxn = get_connection()
    cursor = cnxn.cursor()
    
    #PathSave = os.path.join(os.path.sep)
    PathSave = os.getcwd()

    ruta_del_archivo_existente = os.path.join(BASE_DIR, "Reporte", "ReportesSigreValorizacion.xlsx")
    #POSTES
    cursor.execute(queryValorizacion, CodSed, 'POST')
    records = cursor.fetchall()
    # Obtener columnas
    columns = [column[0] for column in cursor.description]

    Alimentador = []
    Fecha = []
    
    #Fecha	Alimentador	NodoInicial	NodoFinal	Etiqueta	Codigo	BT-109	BT-111	TipoElemento

    data_lists = {
        #"Fecha": [],
        #"Alimentador": [],
        #"NodoInicial": [],
        #"NodoFinal": [],
        "Etiqueta": [],
        "Codigo": [],
        "BT-109":[],
        "Espacio": [],
        "BT-111": [],
        "TipoElemento": []
    }

    TotalBT109 = 0
    TotalBT111 = 0

    # Iterate through records
    for rows in records:
        #data_lists["NodoInicial"].append(str(rows[3]).replace('None','')) 
        #data_lists["NodoFinal"].append(str(rows[2]).replace('None',''))
        data_lists["Etiqueta"].append(str(rows[4]).replace('None',''))
        data_lists["Codigo"].append(str(rows[5]).replace('None',''))
        data_lists["BT-109"].append(str(rows[6]).replace('None',''))
        data_lists["BT-111"].append(str(rows[8]).replace('None',''))
        TotalBT109 += int(str(rows[6]).replace('None','0'))
        TotalBT111 += int(str(rows[8]).replace('None','0'))
        Alimentador.append(str(rows[1]))
        Fecha.append(str(rows[0]))

    # Create a list of results by appending the lists from the dictionary
    results = [data_lists[key] for key in data_lists]
    # Cargar el archivo Excel
    df = pd.DataFrame(results)

    # Paso 2: Cargar el archivo Excel existente
    wb = load_workbook(ruta_del_archivo_existente)

    hoja = wb['BT'] 
    
    print(len(results[2]))
    
    # Paso 3: Seleccionar la hoja en la que deseas escribir
    # Puedes usar wb['Nombre_de_la_Hoja'] si no es la hoja activa

    hoja['N4'] = 'ARJEN SRL'
    hoja['N5'] = Alimentador[0] + ' / ' + CodSed
    hoja['N6'] = Fecha[0]
    hoja['N7'] = NroOrden
    rango_origen_CuadroSumTotal = hoja["B8:I23"] 
    rango_origenColumnas = hoja["M8:M37"]


    for idx in range(1,len(results[2])):
        copiar_formato(hoja,rango_origenColumnas,0,idx)

    # Paso 4: Escribir los datos del DataFrame en la hoja
    for r_idx, row in enumerate(df.iterrows(), 1):
        for c_idx, value in enumerate(row[1], 1):
            celda = hoja.cell(row=r_idx+7, column=c_idx+12, value=value)
            
    
    copiar(hoja,rango_origen_CuadroSumTotal,0,len(results[2]) + 11)
    
    TotalLen = len(results[2])

    hoja.cell(row= 10, column=TotalLen+13, value=TotalBT109)
    hoja.cell(row= 12, column=TotalLen+13, value=TotalBT111)

    # VANOS ---------------------------------------------------------------------------------------------------------------------------------------------------

    cursor.execute(queryValorizacion, CodSed, 'VANO')
    records = cursor.fetchall()

    results = []

    Tram = []
    Codigo = []
    NodoInicial = []
    NodoFinal = []
    DEF = []
    SINDEF = []

    #Alimentador = []
    #Fecha=[]
    TotalBT110 = 0
    TotalBT112 = 0

    for rows in records:
            #Alimentador.append(str(rows[0]).replace('None',''))
            #Fecha.append(str(rows[1]).replace('None',''))
            Codigo.append(str(rows[5]).replace('None',''))
            NodoInicial.append(str(rows[2]).replace('None',''))
            NodoFinal.append(str(rows[3]).replace('None',''))
            DEF.append(str(rows[7]).replace('None',''))
            SINDEF.append(str(rows[9]).replace('None',''))
            TotalBT110 += int(str(rows[7]).replace('None','0'))
            TotalBT112 += int(str(rows[9]).replace('None','0'))


    results.append(NodoInicial)
    results.append(NodoFinal)
    results.append(Codigo)
    results.append([''])
    results.append(DEF)
    results.append([''])
    results.append(SINDEF)

        # Cargar el archivo Excel
    DFVanos = pd.DataFrame(results)

    hoja = wb['VBT']

    print(len(results[2]))

    hoja['N4'] = 'ARJEN SRL'
    hoja['N5'] = Alimentador[0] + ' / ' + CodSed
    hoja['N6'] = Fecha[0]
    hoja['N7'] = NroOrden
    rango_origen_CuadroSumTotal = hoja["B8:I23"] 
    rango_origenColumnas = hoja["M8:M37"]

 
    for idx in range(1,len(results[2])):
        copiar_formato(hoja,rango_origenColumnas,0,idx)

    # Paso 4: Escribir los datos del DataFrame en la hoja
    for r_idx, row in enumerate(DFVanos.iterrows(), 1):
        for c_idx, value in enumerate(row[1], 1):
            celda = hoja.cell(row=r_idx+7, column=c_idx+12, value=value)
            
            
    copiar(hoja,rango_origen_CuadroSumTotal,0,len(results[2]) + 11)
    
    hoja.cell(row= 12, column=len(results[2])+  13, value=TotalBT110)
    hoja.cell(row= 14, column=len(results[2]) + 13, value=TotalBT112)


    #AMBOS ----------------------------------------------------------------------------------------------
    
    
    cursor.execute(queryValorizacion, CodSed, 'AMBOS')
    records = cursor.fetchall()

    results = []

    Codigo = []
    CodigoNodo = []
    TipoElemento = []
    Etiqueta = []
    NodoFinal = []
    BT109 = []
    BT110 = []
    BT111 = []
    BT112 = []


    for rows in records:
            #Alimentador.append(str(rows[0]).replace('None',''))
            #Fecha.append(str(rows[1]).replace('None',''))
            Codigo.append(str(rows[5]).replace('None',''))
            CodigoNodo.append(str(rows[11]).replace('None',''))
            TipoElemento.append(str(rows[10]).replace('None',''))
            Etiqueta.append(str(rows[4]).replace('None',''))
            BT109.append(int(rows[6]) if rows[6] else 0)
            BT110.append(int(rows[7]) if rows[7] else 0)
            BT111.append(int(rows[8]) if rows[8] else 0)
            BT112.append(int(rows[9]) if rows[9] else 0)

    
    results.append(TipoElemento)
    results.append(CodigoNodo)
    results.append(Codigo)
    results.append(BT109)
    results.append(BT110)
    results.append(BT111)
    results.append(BT112)

        # Cargar el archivo Excel
    DFVanos = pd.DataFrame(results)

    hoja = wb['VALORIZADO']

    print(len(results[2]))

    hoja['N4'] = 'ARJEN SRL'
    hoja['N5'] = Alimentador[0] + ' / ' + CodSed
    hoja['N6'] = Fecha[0]
    hoja['N7'] = NroOrden
    rango_origen_CuadroSumTotal = hoja["B8:I23"] 
    rango_origenColumnas = hoja["M8:M37"]

 
    for idx in range(1,len(results[2])):
        copiar_formato(hoja,rango_origenColumnas,0,idx)

    # Paso 4: Escribir los datos del DataFrame en la hoja
    for r_idx, row in enumerate(DFVanos.iterrows(), 1):
        for c_idx, value in enumerate(row[1], 1):
            celda = hoja.cell(row=r_idx+7, column=c_idx+12, value=value)
    
    copiar(hoja,rango_origen_CuadroSumTotal,0,len(results[2]) + 11)
    
    # Crear la fórmula de suma en formato de Excel, por ejemplo, SUM(L7:XX[len(results[2])+13]7)
    # Lo colocamos en la fila 11 (row=11) y columna (len(results[2])+13)
    col_inicio = 13
    col_fin = len(results[2]) + 12
    col_inicio_letra = get_column_letter(col_inicio)
    col_fin_letra = get_column_letter(col_fin)

    formula1 = f"=SUM({col_inicio_letra}11:{col_fin_letra}11)"
    hoja.cell(row=11, column=len(results[2]) + 13, value=formula1)

    formula2 = f"=SUM({col_inicio_letra}12:{col_fin_letra}12)"
    hoja.cell(row= 12, column=len(results[2])+13, value=formula2)

    formula3 = f"=SUM({col_inicio_letra}13:{col_fin_letra}13)"
    hoja.cell(row= 13, column=len(results[2])+13, value=formula3)
    
    formula4 = f"=SUM({col_inicio_letra}14:{col_fin_letra}14)"
    hoja.cell(row= 14, column=len(results[2])+13, value=formula4)


    #cursor.execute("rollback TRANSACTION")
    File = 'Valorizado '+Alimentador[0]+'-'+ CodSed+'.xlsx'

    # Paso 5: Guardar el archivo Excel con otro nombre
    nueva_ruta = os.path.join(PathSave, File)
    #nueva_ruta = 'C:/Users/SIGRE/Desktop/Compartido2/Reportes Generados/Reporte '+Alimentador[0]+' '+datetime.now().strftime("%d%m%Y%H%M")+'.xlsx'
    print("Path:",nueva_ruta)
    wb.save(nueva_ruta)
    
    return nueva_ruta,File
