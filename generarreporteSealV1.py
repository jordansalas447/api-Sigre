from datetime import datetime
import pyodbc
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from io import BytesIO
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, PatternFill
import asyncio
from flask import jsonify
from SigreApiRest.config import get_connection
from SigreApiRest.utils import copiar_formato, Unirceldas, ConvertirNoneto0 , copiar

#cnxn = Config.cnxn
#cursor = cnxn.cursor()

def GenerarReporte(CodAlim,PathSave,BASEPATH):

    cnxn = get_connection()
    cursor = cnxn.cursor()
    
    ruta_del_archivo_existente = r'C:\Users\Usuario\Documents\SigreWeb\sigreweb-main\SigreApiRest\Reporte\ReportesSigre.xlsx'
    #POSTES
    #cursor.execute("BEGIN TRANSACTION")  
    #cursor.execute("sp_UpdateCorregirRutas " + CodAlim)
    cursor.execute("exec sp_GetReportPostsBTByFeeder " + CodAlim)
    records = cursor.fetchall()
    
    # Obtener columnas
    columns = [column[0] for column in cursor.description]

    Alimentador = []
    Fecha = []

    data_lists = {
        "POST_Etiqueta": [],
        "POST_CodigoNodo": [],
        "NEMA": [],
        "TIPI6002": [],
        "TIPI6004": [],
        "TIPI6006": [],
        "TIPI6008":[],
        "Espacio01": [],
        "TIPI6024": [],
        "Tipo": [],
        "Espacio02":"",
        #"TipoArmado": [],
        #"ArmadoMaterial": [],
        #"Espacio03":"",
        #"Espacio04":"",
        "TIPI6026": [],
        "TIPI6028": [],
        "Espacio05":"",
        #"TIPI1082": [],
        #"TIPI1086": [],
        "Criticidad": [],
        "Fotos":[],
        "Ruta": [],
        "Espacio06":"",
        "Espacio07":"",
        "Espacio08":"",
        #"S0": [],
        #"S1": [],
        #"S2": [],
        "N": [],
        "Total": [],
    }

    # Iterate through records
    for rows in records:
        data_lists["POST_Etiqueta"].append(str(rows[3]).replace('None','')) 
        data_lists["POST_CodigoNodo"].append(str(rows[2]).replace('None',''))
        data_lists["NEMA"].append(str(rows[6]).replace('None',''))
        data_lists["TIPI6002"].append(str(rows[7]).replace('None',''))
        data_lists["TIPI6004"].append(str(rows[8]).replace('None',''))
        data_lists["TIPI6006"].append(str(rows[9]).replace('None',''))
        data_lists["TIPI6008"].append(str(rows[10]).replace('None',''))
        data_lists["TIPI6024"].append(str(rows[11]).replace('None',''))
        #data_lists["Tipo"].append(str(rows[12]).replace('None',''))
        #data_lists["TipoArmado"].append(str(rows[4]).replace('None',''))
        #data_lists["ArmadoMaterial"].append(str(rows[5]).replace('None',''))
        data_lists["TIPI6026"].append(str(rows[12]).replace('None',''))
        data_lists["TIPI6028"].append(str(rows[13]).replace('None',''))
        #data_lists["TIPI1082"].append(str(rows[15]).replace('None',''))
        #data_lists["TIPI1086"].append(str(rows[16]).replace('None',''))
        data_lists["Criticidad"].append(str(rows[14]).replace('None',''))
        data_lists['Fotos'].append(str(rows[20]).replace('None',''))
        data_lists["Ruta"].append('=HYPERLINK("'+BASEPATH+  str(rows[15]).replace('None','') + '","Ver Fotos")')
        #data_lists["S0"].append(int(ConvertirNoneto0(rows[16])))
        #data_lists["S1"].append(int(ConvertirNoneto0(rows[17])))
        #data_lists["S2"].append(int(ConvertirNoneto0(rows[18])))
        data_lists["N"].append(int(ConvertirNoneto0(rows[19])))
        data_lists["Total"].append(ConvertirNoneto0(rows[16]) + ConvertirNoneto0(rows[17]) + ConvertirNoneto0(rows[18]) + ConvertirNoneto0(rows[19]))
        
        Alimentador.append(str(rows[0]))
        Fecha.append(str(rows[1]))

    # Create a list of results by appending the lists from the dictionary
    results = [data_lists[key] for key in data_lists]

    # Cargar el archivo Excel
    df = pd.DataFrame(results)

    # Paso 2: Cargar el archivo Excel existente
    wb = load_workbook(ruta_del_archivo_existente)

    hoja = wb['BT'] 
    
    print(len(results[2]))
    
    # Paso 3: Seleccionar la hoja en la que deseas escribir
    # Puedes usar wb['Nombre_de_la_Hoja'] si no es la hoja activa

    hoja['N4'] = 'ARJEN SRL'
    hoja['N5'] = Alimentador[0]
    hoja['N6'] = Fecha[0]
    rango_origen_CuadroSumTotal = hoja["B8:I23"] 
    rango_origenColumnas = hoja["M8:M37"]


    for idx in range(1,len(results[2])):
        copiar_formato(hoja,rango_origenColumnas,0,idx)

    # Paso 4: Escribir los datos del DataFrame en la hoja
    for r_idx, row in enumerate(df.iterrows(), 1):
        for c_idx, value in enumerate(row[1], 1):
            celda = hoja.cell(row=r_idx+7, column=c_idx+12, value=value)
            if value == 'Sí':
                celda.font = Font(bold=True,size=14,name='Arial')

    celda = hoja.cell(row=r_idx-13, column=c_idx+13, value=len(results[2]))
    celda.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
    celda.font = Font(bold=True, color="000000", size=16)  # Letra negra, en negrita, tamaño grande
    from openpyxl.styles import Border, Side  # Asegúrate de haber importado Border y Side antes
    thick_black = Border(
        left=Side(border_style='medium', color='000000'),
        right=Side(border_style='medium', color='000000'),
        top=Side(border_style='medium', color='000000'),
        bottom=Side(border_style='medium', color='000000')
    )
    celda.border = thick_black
    # Definir el rango que quieres copiar

    copiar(hoja,rango_origen_CuadroSumTotal,0,len(results[2]) + 13)

    cursor.execute("exec sp_GetCountDefBTByFeeder " + CodAlim)
    records = cursor.fetchall()
    results2 = []

    #TotalS0=0
    #TotalS1=0
    #TotalS2=0
    TotalN=0
    TotalSuma=0

    for rows in records:
        data = {
        'Codigo':rows[0],
        #'S0':rows[1],
        #'S1':rows[2],
        #'S2':rows[3],
        'N':rows[4],
        'Total':int(rows[1])+int(rows[2])+int(rows[3])+int(rows[4]),
        'Suma':int(rows[1])+int(rows[4])
        }
        results2.append(data)

    for list in results2:
        #TotalS0 += list['S0']
        #TotalS1 += list['S1']
        #TotalS2 += list['S2']
        TotalN += list['N']
        TotalSuma += list['Suma']

    data = {
        'Codigo':'Total',
        #'S0':TotalS0,
        #'S1':TotalS1,
        #'S2':TotalS2,
        'N':TotalN,
        'Total':'',
        'Suma':TotalSuma
        }
    results2.append(data) 

    df2 = pd.DataFrame(results2)


    # Cuadro final de suma
    for row_idx, row_data in df2.iterrows():
        for col_idx, col_name in enumerate(df2.columns, 1):
            cell = hoja.cell(row=row_idx + 13, column=col_idx +len(results[2]) + 14).value = row_data[col_name]


    Unirceldas(hoja,len(results[2])+14,8,len(results[2])+14,12)
    #Unirceldas(hoja,len(results[2])+15,8,len(results[2])+17,11)
    Unirceldas(hoja,len(results[2])+15,8,len(results[2])+15,11)
    Unirceldas(hoja,len(results[2])+16,8,len(results[2])+16,12)
    Unirceldas(hoja,len(results[2])+17,8,len(results[2])+17,12)


    # VANOS ---------------------------------------------------------------------------------------------------------------------------------------------------

    cursor.execute("exec sp_GetReportGapsVBTByFeeder " + CodAlim)
    records = cursor.fetchall()

    results = []

    Tram = []
    VANO_Codigo = []
    NodoInicial = []
    NodoFinal = []
    TIPI7002 = []
    TIPI7004 = []
    TIPI7006 = []
    TIPI7008 = []
    TIPI5030 = []
    TIPI5032 = []
    TIPI5038 = []
    Fotos = []
    Criticidad = []
    Ruta = []
    

    #S0 = []
    #S1 = []
    #S2 = []
    N = []

    Alimentador = []
    Fecha=[]
    Total = []

    for rows in records:
            Alimentador.append(str(rows[0]).replace('None',''))
            Fecha.append(str(rows[1]).replace('None',''))
            VANO_Codigo.append(str(rows[2]).replace('None',''))
            NodoInicial.append(str(rows[3]).replace('None',''))
            NodoFinal.append(str(rows[4]).replace('None','')),
            TIPI7002.append(str(rows[5]).replace('None',''))
            TIPI7004.append(str(rows[6]).replace('None',''))
            TIPI7006.append(str(rows[7]).replace('None',''))
            TIPI7008.append(str(rows[8]).replace('None',''))
            Criticidad.append(str(rows[9]).replace('None',''))
            Fotos.append(str(rows[15]).replace('None',''))
            Ruta.append('=HYPERLINK("'+BASEPATH+  str(rows[10]).replace('None','') + '","Ver Fotos")')

            #S0.append(int(ConvertirNoneto0(rows[11])))
            #S1.append(int(ConvertirNoneto0(rows[12])))
            #S2.append(int(ConvertirNoneto0(rows[13])))
            N.append(int(ConvertirNoneto0(rows[14])))

            Total.append(ConvertirNoneto0(rows[11]) + ConvertirNoneto0(rows[12])+ ConvertirNoneto0(rows[13])+ ConvertirNoneto0(rows[14]))          
            #Tram.append(str(rows[15]))


    results.append(Tram)
    results.append(NodoInicial)
    results.append(NodoFinal)
    results.append(VANO_Codigo)
    results.append(TIPI7002)
    results.append(TIPI7004)
    results.append(TIPI7006)
    results.append(TIPI7008)
    #results.append(TIPI5030)
    #results.append(TIPI5032)
    #results.append(TIPI5038)
    results.append(Criticidad)
    results.append(Fotos)
    results.append(Ruta)
    results.append([''])
    results.append([''])
    results.append([''])
    #results.append(S0)
    #results.append(S1)
    #results.append(S2)
    results.append(N)
    results.append(Total)


        # Cargar el archivo Excel
    DFVanos = pd.DataFrame(results)

    hoja = wb['VBT']

    print(len(results[2]))

    # Paso 3: Seleccionar la hoja en la que deseas escribir
     # Puedes usar wb['Nombre_de_la_Hoja'] si no es la hoja activa

    hoja['N4'] = 'ARJEN SRL'
    hoja['N5'] = Alimentador[0]
    hoja['N6'] = Fecha[0]
    rango_origen_CuadroSumTotal = hoja["B6:I19"] 
    rango_origenColumnas = hoja["M8:M37"]


    for idx in range(1,len(results[2])):
        copiar_formato(hoja,rango_origenColumnas,0,idx)

    # Paso 4: Escribir los datos del DataFrame en la hoja
    for r_idx, row in enumerate(DFVanos.iterrows(), 1):
        for c_idx, value in enumerate(row[1], 1):
            celda = hoja.cell(row=r_idx+7, column=c_idx+12, value=value)
            if value == 'Sí':
                celda.font = Font(bold=True,size=14,name='Arial')

    celda = hoja.cell(row=r_idx-5, column=c_idx+13, value=len(results[2]))
    celda.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
    celda.font = Font(bold=True, color="000000", size=16)  # Letra negra, en negrita, tamaño grande
    from openpyxl.styles import Border, Side  # Asegúrate de haber importado Border y Side antes
    thick_black = Border(
        left=Side(border_style='medium', color='000000'),
        right=Side(border_style='medium', color='000000'),
        top=Side(border_style='medium', color='000000'),
        bottom=Side(border_style='medium', color='000000')
    )
    celda.border = thick_black



    # Definir el rango que quieres copiar

    copiar(hoja,rango_origen_CuadroSumTotal,0,len(results[2]) + 13)

    cursor.execute("sp_GetCountDefVBTByFeeder " + CodAlim)
    records = cursor.fetchall()
    results3 = []


    TotalS0=0
    TotalS1=0
    TotalS2=0
    TotalN=0
    TotalSuma=0

    for rows in records:
        data = {
        'Codigo':rows[0],
        #'S0':rows[1],
        #'S1':rows[2],
        #'S2':rows[3],
        'N':rows[4],
        'Total':int(rows[1])+int(rows[2])+int(rows[3])+int(rows[4]),
        'Suma':int(rows[1])+int(rows[4])
        }
        results3.append(data)

    for list in results3:
        #TotalS0 += list['S0']
        #TotalS1 += list['S1']
        #TotalS2 += list['S2']
        TotalN += list['N']
        TotalSuma += list['Suma']

    data = {
        'Codigo':'Total',
        #'S0':TotalS0,
        #'S1':TotalS1,
        #'S2':TotalS2,
        'N':TotalN,
        'Total':'',
        'Suma':TotalSuma
        }
    results3.append(data) 

    DFVAnosCuadroTotal = pd.DataFrame(results3)


    # Cuadro final de suma
    for row_idx, row_data in DFVAnosCuadroTotal.iterrows():
        for col_idx, col_name in enumerate(DFVAnosCuadroTotal.columns, 1):
            cell = hoja.cell(row=row_idx + 12, column=col_idx +len(results[2]) + 14) 
            cell.value = row_data[col_name]

    Unirceldas(hoja,len(results[2])+14,6,len(results[2])+14,11)
    #Unirceldas(hoja,len(results[2])+15,6,len(results[2])+17,10)
    Unirceldas(hoja,len(results[2])+15,6,len(results[2])+15,10)
    Unirceldas(hoja,len(results[2])+16,6,len(results[2])+16,11)
    Unirceldas(hoja,len(results[2])+17,6,len(results[2])+17,11)

    #cursor.execute("rollback TRANSACTION")
    File = Alimentador[0]+' '+datetime.now().strftime("%d%m%Y%H%M")+'.xlsx'

    # Paso 5: Guardar el archivo Excel con otro nombre
    nueva_ruta = PathSave+File
    #nueva_ruta = 'C:/Users/SIGRE/Desktop/Compartido2/Reportes Generados/Reporte '+Alimentador[0]+' '+datetime.now().strftime("%d%m%Y%H%M")+'.xlsx'
    wb.save(nueva_ruta)

    return nueva_ruta,File
